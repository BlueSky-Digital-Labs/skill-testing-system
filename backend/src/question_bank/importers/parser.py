"""
Parse CSV and XLSX spreadsheets into normalized row dictionaries.
"""

from __future__ import annotations

import csv
import io
from typing import BinaryIO

from openpyxl import load_workbook

SUPPORTED_FORMATS = {'csv', 'xlsx'}


def detect_format(filename: str) -> str:
    """
    Determine spreadsheet format from the file name extension.
    """
    if not filename or '.' not in filename:
        raise ValueError('Filename must include a .csv or .xlsx extension.')

    extension = filename.rsplit('.', 1)[-1].lower()
    if extension not in SUPPORTED_FORMATS:
        raise ValueError(
            f'Unsupported file type ".{extension}". Upload a .csv or .xlsx file.',
        )
    return extension


def parse_spreadsheet(file: BinaryIO, filename: str) -> list[dict]:
    """
    Parse spreadsheet content into a list of row dictionaries.

    Keys are normalized to lowercase stripped header names. Empty rows are
    skipped. Each row includes a ``_row_number`` key (1-based, including header).
    """
    file_format = detect_format(filename)
    if file_format == 'csv':
        return _parse_csv(file)
    return _parse_xlsx(file)


def _normalize_header(header: object) -> str:
    return str(header or '').strip().lower()


def _normalize_row(raw_row: dict[str, object], row_number: int) -> dict:
    row = {
        _normalize_header(key): '' if value is None else str(value).strip()
        for key, value in raw_row.items()
        if _normalize_header(key)
    }
    row['_row_number'] = row_number
    return row


def _is_blank_row(row: dict) -> bool:
    return not any(
        value
        for key, value in row.items()
        if key != '_row_number'
    )


def _parse_csv(file: BinaryIO) -> list[dict]:
    content = file.read()
    if isinstance(content, bytes):
        text = content.decode('utf-8-sig')
    else:
        text = content

    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise ValueError('CSV file is missing a header row.')

    rows: list[dict] = []
    for index, raw_row in enumerate(reader, start=2):
        row = _normalize_row(raw_row, index)
        if not _is_blank_row(row):
            rows.append(row)
    return rows


def _parse_xlsx(file: BinaryIO) -> list[dict]:
    workbook = load_workbook(file, read_only=True, data_only=True)
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)

    try:
        header_row = next(row_iter)
    except StopIteration as exc:
        raise ValueError('XLSX file is empty.') from exc

    headers = [_normalize_header(value) for value in header_row]
    if not any(headers):
        raise ValueError('XLSX file is missing a header row.')

    rows: list[dict] = []
    for index, values in enumerate(row_iter, start=2):
        raw_row = {
            headers[col_index]: values[col_index] if col_index < len(values) else None
            for col_index, header in enumerate(headers)
            if header
        }
        row = _normalize_row(raw_row, index)
        if not _is_blank_row(row):
            rows.append(row)
    return rows
